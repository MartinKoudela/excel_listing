import openpyxl

wb = openpyxl.load_workbook('listing_excel.xlsx')
ws = wb.active

print('Zaměstnanec:  ' + ws['A2'].value + '\nOdpracoval:  ' + str(ws['B2'].value) +' Hodin' + '\nV oddělení:  ' + ws['C2'].value)

row_position = 2
col_position = 2

total_hours = ((int(ws.cell(row=row_position, column=col_position).value))+
               (int(ws.cell(row=row_position+1, column=col_position).value))+
               (int(ws.cell(row=row_position+2, column=col_position).value))+
               (int(ws.cell(row=row_position+3, column=col_position).value)))
ws.cell(row=2,column=4).value=total_hours
wb.save('listing_excel.xlsx')